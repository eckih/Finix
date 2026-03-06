"""
US-, Deutschland-, Österreich-, Kanada-, Mexiko-, Schweiz- und Liechtenstein-Finanzdaten:
- USA: TGA-Saldo (täglich) per API; optional FRED: WDTGAL (TGA Mittwoch), RRPONTSYD (Overnight Reverse Repos), WRESBAL (Reserve Balances)
- DE: Kreditbestand Bund (monatlich) aus BMF-Datenportal
- AT: Nettofinanzierungssaldo aus Monatsbericht (BMF Österreich)
- CA: Daily Cash Balance (CRF an der Bank of Canada) per CSV
- MX: Deuda Neta (Schuldenstand) aus SHCP/repodatos
- CH: Finanzierungssaldo/Schulden Bund aus EFV (opendata.swiss / data.finance.admin.ch)
- LI: Keine öffentliche API/CSV; Hinweis auf opendata.li und Regierungs-PDFs
"""
import csv
import io
import logging
import os
import re
import requests

log = logging.getLogger(__name__)

# Optional: .env im Projektroot laden (für FRED_API_KEY beim direkten Aufruf)
try:
    from dotenv import load_dotenv
    _root = os.path.dirname(os.path.abspath(__file__))
    load_dotenv(os.path.join(_root, ".env"))
except ImportError:
    pass

# FRED API (St. Louis Fed) – kostenloser API-Key: https://fred.stlouisfed.org/docs/api/api_key.html
FRED_OBS_URL = "https://api.stlouisfed.org/fred/series/observations"


def _get_fred_latest(series_id: str, api_key: str):
    """Holt den neuesten Beobachtungswert einer FRED-Serie. Liefert dict mit date, value oder None."""
    try:
        r = requests.get(
            FRED_OBS_URL,
            params={
                "series_id": series_id,
                "api_key": api_key,
                "file_type": "json",
                "sort_order": "desc",
                "limit": 1,
            },
            timeout=10,
        )
        r.raise_for_status()
        data = r.json()
        obs = (data.get("observations") or [])
        if not obs:
            return None
        o = obs[0]
        date = o.get("date")
        raw = o.get("value", ".")
        if raw in (None, "", "."):
            return None
        try:
            value = float(raw)
        except (TypeError, ValueError):
            return None
        return {"date": date, "value": value}
    except Exception:
        return None


def _get_fred_observations(series_id: str, api_key: str, limit: int = 100, observation_start: str = None):
    """Holt bis zu `limit` Beobachtungen einer FRED-Serie (neueste zuerst). observation_start z. B. '2020-01-01'."""
    params = {
        "series_id": series_id,
        "api_key": api_key,
        "file_type": "json",
        "sort_order": "desc",
        "limit": min(limit, 10000),
    }
    if observation_start:
        params["observation_start"] = observation_start
    try:
        r = requests.get(
            FRED_OBS_URL,
            params=params,
            timeout=30,
        )
        r.raise_for_status()
        data = r.json()
        obs = data.get("observations") or []
        out = []
        for o in obs:
            date = o.get("date")
            raw = o.get("value", ".")
            if raw in (None, "", "."):
                continue
            try:
                value = float(raw)
            except (TypeError, ValueError):
                continue
            out.append({"date": date, "value": value})
        return out
    except Exception:
        return []

# BMF-URL für Kreditbestand (monatlich, XLSX)
BMF_KREDITBESTAND_XLSX = (
    "https://www.bundesfinanzministerium.de/Datenportal/Daten/offene-daten/"
    "haushalt-oeffentliche-finanzen/Zeitreihe-Kredit-Bruttokredit-Tilgung-Zinsen/"
    "datensaetze/xlsx-Kreditbestand-Bruttokredit-Tilgung-Zinsen.xlsx?__blob=publicationFile&v=28"
)


def get_us_account_balance_pro():
    # Basis-URL der Fiscal Data API (Endpoint: operating_cash_balance, nicht mehr dts_table_1)
    url = "https://api.fiscaldata.treasury.gov/services/api/fiscal_service/v1/accounting/dts/operating_cash_balance"

    # Filter: TGA Closing Balance (Saldo); seit Apr 2022 steht der Wert in open_today_bal
    params = {
        "filter": "account_type:eq:Treasury General Account (TGA) Closing Balance",
        "sort": "-record_date",
        "page[size]": 1
    }

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) IT-Expert-Query/1.0"
    }

    try:
        response = requests.get(url, params=params, headers=headers, timeout=10)
        response.raise_for_status()

        payload = response.json()
        data_list = payload.get("data", [])

        if data_list:
            entry = data_list[0]
            date = entry["record_date"]
            # Ab Apr 2022: Schließsaldo steht in open_today_bal, close_today_bal ist null (API liefert "null" als String)
            def _num(val):
                if val is None or (isinstance(val, str) and val.lower() == "null"):
                    return None
                return float(val)
            raw = _num(entry.get("close_today_bal")) or _num(entry.get("open_today_bal"))
            if raw is None:
                print("Keine Saldo-Werte für den TGA in den aktuellen Daten.")
                return
            balance = raw

            print("--- Offizieller US-Finanzstatus ---")
            print(f"Datum:  {date}")
            print(f"Saldo:  ${balance:,.2f} Mio. USD")
            print(f"Total:  ${(balance / 1000):,.2f} Mrd. USD")
            result = {
                "country": "us",
                "date": date,
                "value": balance / 1000,
                "value_mio": balance,
                "unit": "Mrd. USD",
                "label": "TGA Closing Balance",
            }
            # Optional: FRED-Daten (WDTGAL, RRPONTSYD, WRESBAL, SOFR, EFFR)
            fred_key = os.environ.get("FRED_API_KEY", "").strip()
            if fred_key:
                wdtgal = _get_fred_latest("WDTGAL", fred_key)  # Millions of USD
                rrp = _get_fred_latest("RRPONTSYD", fred_key)  # Billions of USD
                wresbal = _get_fred_latest("WRESBAL", fred_key)  # Millions of USD
                sofr = _get_fred_latest("SOFR", fred_key)  # Percent
                effr = _get_fred_latest("EFFR", fred_key)  # Percent
                if wdtgal:
                    result["fred_wdtgal_date"] = wdtgal["date"]
                    result["fred_wdtgal_value_mio"] = wdtgal["value"]
                    print(f"FRED WDTGAL (TGA Wed): {wdtgal['date']} = ${wdtgal['value']:,.0f} Mio. USD")
                if rrp:
                    result["fred_rrpontsyd_date"] = rrp["date"]
                    result["fred_rrpontsyd_value_mrd"] = rrp["value"]
                    print(f"FRED RRPONTSYD (Overnight RRP): {rrp['date']} = ${rrp['value']:,.3f} Mrd. USD")
                if wresbal:
                    result["fred_wresbal_date"] = wresbal["date"]
                    result["fred_wresbal_value_mio"] = wresbal["value"]
                    print(f"FRED WRESBAL (Reserve Balances): {wresbal['date']} = ${wresbal['value']:,.0f} Mio. USD")
                if sofr:
                    result["fred_sofr_date"] = sofr["date"]
                    result["fred_sofr_value"] = sofr["value"]
                    print(f"FRED SOFR: {sofr['date']} = {sofr['value']:.2f} %")
                if effr:
                    result["fred_effr_date"] = effr["date"]
                    result["fred_effr_value"] = effr["value"]
                    print(f"FRED EFFR: {effr['date']} = {effr['value']:.2f} %")
            return result
        else:
            print("Keine aktuellen Daten in der Tabelle gefunden.")
            return None

    except requests.exceptions.RequestException as e:
        print(f"Netzwerk- oder API-Fehler: {e}")
        return None
    except ValueError as e:
        print(f"Fehler beim Verarbeiten der Daten: {e}")
        return None


def get_us_historical(limit_treasury: int = 100, limit_fred: int = 100, start_date: str = None):
    """
    Holt historische Daten für TGA (Treasury), WDTGAL, RRPONTSYD, WRESBAL, SOFR und EFFR (FRED).
    start_date z. B. '2020-01-01' – dann werden Daten ab diesem Datum geholt.
    Liefert eine Liste von Datensätzen zum Speichern (country, date, value, unit, label).
    """
    records = []
    start_date = (start_date or "2020-01-01").strip() or None
    url = "https://api.fiscaldata.treasury.gov/services/api/fiscal_service/v1/accounting/dts/operating_cash_balance"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) IT-Expert-Query/1.0"}

    def _num(val):
        if val is None or (isinstance(val, str) and val.lower() == "null"):
            return None
        return float(val)

    try:
        page = 1
        page_size = 1000
        max_pages = 30
        while page <= max_pages:
            params = {
                "filter": "account_type:eq:Treasury General Account (TGA) Closing Balance",
                "sort": "-record_date",
                "page[size]": page_size,
                "page[number]": page,
            }
            resp = requests.get(url, params=params, headers=headers, timeout=20)
            resp.raise_for_status()
            data_list = resp.json().get("data", [])
            if not data_list:
                break
            oldest_in_page = None
            for entry in data_list:
                date = entry.get("record_date")
                if date:
                    oldest_in_page = date if oldest_in_page is None else min(oldest_in_page, date)
                if start_date and date and date < start_date:
                    continue
                raw = _num(entry.get("close_today_bal")) or _num(entry.get("open_today_bal"))
                if date and raw is not None:
                    records.append({
                        "country": "us",
                        "date": date,
                        "value": raw / 1000.0,
                        "unit": "Mrd. USD",
                        "label": "TGA Closing Balance",
                    })
            if start_date and oldest_in_page is not None and oldest_in_page < start_date:
                break
            if len(data_list) < page_size:
                break
            page += 1
    except Exception:
        pass

    fred_key = os.environ.get("FRED_API_KEY", "").strip()
    if fred_key:
        limit_fred_actual = max(limit_fred, 2500) if start_date else limit_fred
        for series_id, label, unit in [
            ("WDTGAL", "WDTGAL (TGA Wed)", "Mrd. USD"),
            ("RRPONTSYD", "RRPONTSYD (Overnight RRP)", "Mrd. USD"),
            ("WRESBAL", "WRESBAL (Reserve Balances)", "Mrd. USD"),
            ("SOFR", "SOFR", "%"),
            ("EFFR", "EFFR", "%"),
        ]:
            obs = _get_fred_observations(
                series_id, fred_key, limit=limit_fred_actual,
                observation_start=start_date,
            )
            # Fallback: FRED kann bei Serien-ID Groß-/Kleinschreibung sensibel sein
            if not obs and series_id == "WRESBAL":
                obs = _get_fred_observations(
                    "wresbal", fred_key, limit=limit_fred_actual,
                    observation_start=start_date,
                )
            if series_id == "WRESBAL" and not obs:
                log.warning("FRED WRESBAL: keine Daten (series_id=WRESBAL und wresbal). Bitte FRED_API_KEY prüfen.")
            if obs:
                scale = 1.0 / 1000.0 if series_id in ("WDTGAL", "WRESBAL") else 1.0
                for o in obs:
                    records.append({
                        "country": "us",
                        "date": o["date"],
                        "value": o["value"] * scale,
                        "unit": unit,
                        "label": label,
                    })
    return records


def get_markets_historical(limit_fred: int = 500, start_date: str = None):
    """
    Holt historische Kursdaten von FRED: S&P 500 (SP500) und Bitcoin (CBBTCUSD).
    Speicherung unter country="markets" für die Kurse-Ansicht.
    """
    records = []
    start_date = (start_date or "2020-01-01").strip() or None
    fred_key = os.environ.get("FRED_API_KEY", "").strip()
    if not fred_key:
        return records
    limit_actual = max(limit_fred, 2500) if start_date else limit_fred
    for series_id, label, unit in [
        ("SP500", "S&P 500", "Index"),
        ("CBBTCUSD", "BTC", "USD"),
    ]:
        obs = _get_fred_observations(
            series_id, fred_key, limit=limit_actual,
            observation_start=start_date,
        )
        for o in obs:
            records.append({
                "country": "markets",
                "date": o["date"],
                "value": o["value"],
                "unit": unit,
                "label": label,
            })
    return records


def get_de_account_balance():
    """
    Kreditbestand (Schuldenstand) des Bundes aus dem BMF-Datenportal.
    Deutschland veröffentlicht keinen täglichen Kontostand wie die USA;
    es wird der monatliche Kreditbestand aus der XLSX-Zeitreihe gelesen.
    """
    try:
        import io
        import openpyxl
    except ImportError:
        print("Für Deutschland-Abfrage bitte installieren: pip install openpyxl")
        return

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) IT-Expert-Query/1.0"
    }
    try:
        response = requests.get(
            BMF_KREDITBESTAND_XLSX, headers=headers, timeout=25
        )
        response.raise_for_status()
        wb = openpyxl.load_workbook(
            io.BytesIO(response.content), read_only=True, data_only=True
        )
        sheet_name = "rpgSchuldenstand"
        if sheet_name not in wb.sheetnames:
            print("Erwartetes Arbeitsblatt in der BMF-Datei nicht gefunden.")
            wb.close()
            return
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Zeile 2 (Index 2): Datumswerte ab Spalte 1; Zeile 3: Kreditbestand
        if len(rows) < 4:
            print("Unerwartetes Format der BMF-Datei.")
            return
        date_row = list(rows[2])
        value_row = list(rows[3])
        # Letztes gültiges Datum und zugehöriger Wert
        last_date = None
        last_value = None
        for i in range(1, min(len(date_row), len(value_row))):
            d, v = date_row[i], value_row[i]
            if d is not None and v is not None:
                try:
                    _ = float(v)
                    last_date, last_value = d, float(v)
                except (TypeError, ValueError):
                    pass
        if last_date is None or last_value is None:
            print("Keine gültigen Kreditbestand-Daten in der BMF-Datei.")
            return
        # Werte in der Datei sind in vollen Euro; Ausgabe in Mrd. EUR
        value_mrd = last_value / 1_000_000_000
        print("--- Offizieller Finanzstatus Bund (BMF) ---")
        print(f"Datum:  {last_date}")
        print(f"Kreditbestand (Schuldenstand):  {value_mrd:,.2f} Mrd. EUR")
        return {
            "country": "de",
            "date": str(last_date),
            "value": value_mrd,
            "unit": "Mrd. EUR",
            "label": "Kreditbestand (Schuldenstand)",
        }
    except requests.exceptions.RequestException as e:
        print(f"Netzwerk- oder API-Fehler: {e}")
        return None
    except Exception as e:
        print(f"Fehler beim Verarbeiten der Daten: {e}")
        return None


# Basis-URL für österreichischen Monatsbericht (Excel-Link wird von der Übersichtsseite geholt)
BMF_AT_OVERVIEW = "https://www.bmf.gv.at/services/startseite-budget/Monatliche-Berichterstattung/Monatlicher-Vollzug-2026.html"

# Kanada: Daily Cash Balance (Consolidated Revenue Fund), Receiver General
CA_DAILY_CASH_CSV = "https://donnees-data.tpsgc-pwgsc.gc.ca/ba1/sqt-dcb/sqt-dcb.csv"

# Mexiko: Deuda pública (SHCP Estadísticas oportunas), repodatos
MX_DEUDA_PUBLICA_CSV = "https://repodatos.atdt.gob.mx/s_hacienda_cred_publico/indicadores_fiscales/deuda_publica.csv"

# Schweiz: Hauptaggregate EFV (Bundesfinanzen), data.finance.admin.ch
CH_EFV_MAIN_CSV = "https://www.data.finance.admin.ch/static/assets/datasets/fs_dashboard/main_extern.csv"


def get_at_account_balance():
    """
    Nettofinanzierungssaldo (Finanzierungshaushalt) aus dem Monatsbericht
    des österreichischen BMF. Kein täglicher Kontostand; es wird der
    Monatserfolg aus der aktuellen Excel-Tabelle gelesen.
    """
    try:
        import io
        import openpyxl
    except ImportError:
        print("Für Österreich-Abfrage bitte installieren: pip install openpyxl")
        return None

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) IT-Expert-Query/1.0"
    }
    try:
        # Aktuellen Excel-Link von der Übersichtsseite holen
        overview = requests.get(BMF_AT_OVERVIEW, headers=headers, timeout=15)
        overview.raise_for_status()
        text = overview.text
        # Suche nach Link zur Tabellen-XLSX (z.B. .../Monatsbericht_Jänner_2026_Tabellen.xlsx)
        match = re.search(r'href="([^"]*Tabellen\.xlsx)[^"]*"', text, re.IGNORECASE)
        if match:
            excel_url = match.group(1).strip()
            if excel_url.startswith("/"):
                excel_url = "https://www.bmf.gv.at" + excel_url
            elif not excel_url.startswith("http"):
                excel_url = "https://www.bmf.gv.at/" + excel_url.lstrip("/")
        else:
            # Fallback: bekannter Link (kann monatlich veraltet sein)
            excel_url = (
                "https://www.bmf.gv.at/dam/jcr:eebb0fc8-8256-466e-a7d5-b08045919ca7/"
                "Monatsbericht_J%C3%A4nner_2026_Tabellen.xlsx"
            )

        response = requests.get(excel_url, headers=headers, timeout=25)
        response.raise_for_status()
        wb = openpyxl.load_workbook(
            io.BytesIO(response.content), read_only=True, data_only=True
        )
        if "T1" not in wb.sheetnames:
            print("Erwartetes Arbeitsblatt T1 im österreichischen Monatsbericht nicht gefunden.")
            wb.close()
            return None
        ws = wb["T1"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # T1: Zeile mit "Nettofinanzierungssaldo", Monatserfolg in Spalte 7; Monat/Jahr aus Zeile 2/3
        month_label = None
        net_saldo = None
        for i, row in enumerate(rows):
            row = list(row) if row else []
            if len(row) > 0 and row[0] and "Nettofinanzierungssaldo" in str(row[0]):
                if len(row) > 7 and row[7] is not None:
                    try:
                        net_saldo = float(row[7])
                    except (TypeError, ValueError):
                        pass
                break
        if len(rows) > 2:
            r2 = list(rows[2]) if len(rows) > 2 else []
            r3 = list(rows[3]) if len(rows) > 3 else []
            if len(r2) > 7 and r2[7]:
                month_label = str(r2[7]).strip()
            if len(r3) > 7 and r3[7]:
                year = r3[7]
                if month_label and year is not None:
                    month_label = f"{month_label} {year}"

        if net_saldo is None:
            print("Keine Nettofinanzierungssaldo-Daten im österreichischen Monatsbericht gefunden.")
            return None
        print("--- Offizieller Finanzstatus Bund (BMF Österreich) ---")
        if month_label:
            print(f"Berichtsmonat:  {month_label}")
        print(f"Nettofinanzierungssaldo (Monatserfolg):  {net_saldo:,.2f} Mio. EUR")
        print(f"  (= {net_saldo / 1000:,.2f} Mrd. EUR)")
        return {
            "country": "at",
            "date": month_label or "",
            "value": net_saldo / 1000,
            "value_mio": net_saldo,
            "unit": "Mrd. EUR",
            "label": "Nettofinanzierungssaldo",
        }
    except requests.exceptions.RequestException as e:
        print(f"Netzwerk- oder API-Fehler: {e}")
        return None
    except Exception as e:
        print(f"Fehler beim Verarbeiten der Daten: {e}")
        return None


def get_ca_account_balance():
    """
    Täglicher Kassenstand (Daily Cash Balance) des Consolidated Revenue Fund (CRF)
    an der Bank of Canada. Daten: Receiver General / open.canada.ca, CSV (quartalsweise aktualisiert).
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) IT-Expert-Query/1.0"
    }
    try:
        try:
            response = requests.get(
                CA_DAILY_CASH_CSV, headers=headers, timeout=25, verify=True
            )
        except requests.exceptions.SSLError:
            import urllib3
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)  # nur bei Fallback
            response = requests.get(
                CA_DAILY_CASH_CSV, headers=headers, timeout=25, verify=False
            )
            print("Hinweis: SSL-Zertifikat konnte nicht geprüft werden (verify=False).")
        response.raise_for_status()
        text = response.text
        if not text.strip():
            print("Kanada: CSV-Datei ist leer.")
            return None
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if len(rows) < 2:
            print("Kanada: Keine Datenzeilen in der CSV.")
            return None
        header = [str(h).strip().lower() for h in rows[0]]
        # Typische Spalten: Date, Balance / Closing Balance / Solde de clôture
        date_col = None
        balance_col = None
        for i, h in enumerate(header):
            if not h:
                continue
            if "date" in h or "dat" in h:
                date_col = i
            if "balance" in h or "solde" in h or "closing" in h:
                balance_col = i
        if date_col is None:
            date_col = 0
        if balance_col is None:
            balance_col = 1 if len(header) > 1 else 0
        # Letzte Datenzeile (evtl. Leerzeilen am Ende überspringen)
        data_rows = [r for r in rows[1:] if len(r) > max(date_col, balance_col) and r[date_col]]
        if not data_rows:
            print("Kanada: Keine gültigen Datensätze in der CSV.")
            return None
        last = data_rows[-1]
        date_val = last[date_col].strip()
        raw_balance = last[balance_col].strip().replace(",", "").replace(" ", "")
        try:
            balance = float(raw_balance)
        except ValueError:
            print("Kanada: Saldo-Wert konnte nicht gelesen werden.")
            return None
        print("--- Offizieller Kassenstand Kanada (Receiver General) ---")
        print(f"Datum:  {date_val}")
        if abs(balance) >= 1_000_000_000:
            value_mrd = balance / 1_000_000_000
            print(f"Saldo (CRF):  ${value_mrd:,.2f} Mrd. CAD")
            return {
                "country": "ca",
                "date": date_val,
                "value": value_mrd,
                "unit": "Mrd. CAD",
                "label": "Daily Cash Balance (CRF)",
            }
        print(f"Saldo (CRF):  ${balance:,.2f} CAD")
        return {
            "country": "ca",
            "date": date_val,
            "value": balance,
            "unit": "CAD",
            "label": "Daily Cash Balance (CRF)",
        }
    except requests.exceptions.RequestException as e:
        print(f"Netzwerk- oder API-Fehler: {e}")
        return None
    except Exception as e:
        print(f"Fehler beim Verarbeiten der Daten: {e}")
        return None


# Monatsnamen für Sortierung (Spanisch)
_MES_ORDER = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]


def get_mx_account_balance():
    """
    Deuda Neta (Schuldenstand) des Sector Público Federal aus den
    Estadísticas oportunas de finanzas públicas (SHCP), CSV über repodatos.atdt.gob.mx.
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) IT-Expert-Query/1.0"
    }
    try:
        response = requests.get(
            MX_DEUDA_PUBLICA_CSV, headers=headers, timeout=35
        )
        response.raise_for_status()
        # CSV oft ISO-8859-1 („Público“); sonst Zeichen können falsch matchen
        try:
            text = response.content.decode("utf-8")
        except UnicodeDecodeError:
            text = response.content.decode("iso-8859-1")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            print("Mexiko: Keine Daten in der CSV.")
            return None
        # Deuda Neta in Millionen Pesos (Sector Público Federal)
        def is_deuda_neta_mxn(row):
            n = (row.get("nombre") or "").upper()
            return "NETA" in n and "PESOS" in n
        federal = [r for r in rows if is_deuda_neta_mxn(r)]
        # Bevorzuge Sector Público Federal (enthält „Federal“ unabhängig von Encoding)
        federal = [r for r in federal if "Federal" in (r.get("sector") or "")] or federal
        if not federal:
            print("Mexiko: Kein Indikator 'Deuda Neta' (Millones de pesos) gefunden.")
            return None
        # Neueste Periode (ciclo, mes)
        def period_key(r):
            ciclo = r.get("ciclo", "0")
            mes = r.get("mes", "")
            try:
                mes_i = _MES_ORDER.index(mes) if mes in _MES_ORDER else -1
            except ValueError:
                mes_i = -1
            return (int(ciclo) if ciclo.isdigit() else 0, mes_i)
        latest = max(federal, key=period_key)
        ciclo = latest.get("ciclo", "")
        mes = latest.get("mes", "")
        raw = (latest.get("monto") or "").strip().replace(",", "")
        try:
            monto_millones = float(raw)
        except ValueError:
            print("Mexiko: Saldo-Wert konnte nicht gelesen werden.")
            return None
        monto_mrd = monto_millones / 1_000
        print("--- Offizieller Schuldenstand Mexiko (SHCP) ---")
        print(f"Periode:  {mes} {ciclo}")
        print(f"Deuda Neta (Sector Público Federal):  {monto_millones:,.0f} Mio. MXN")
        if monto_mrd >= 1_000:
            print(f"  (= {monto_mrd / 1_000:,.2f} Bio. MXN)")
        else:
            print(f"  (= {monto_mrd:,.2f} Mrd. MXN)")
        return {
            "country": "mx",
            "date": f"{mes} {ciclo}",
            "value": monto_mrd,
            "unit": "Mrd. MXN",
            "label": "Deuda Neta (Sector Público Federal)",
        }
    except requests.exceptions.RequestException as e:
        print(f"Netzwerk- oder API-Fehler: {e}")
        return None
    except Exception as e:
        print(f"Fehler beim Verarbeiten der Daten: {e}")
        return None


def get_ch_account_balance():
    """
    Finanzierungssaldo und optional Bruttoschuld des Bundes (Staat/Konföderation)
    aus den Hauptaggregaten der EFV (data.finance.admin.ch / opendata.swiss).
    Werte in Mio. CHF, jährlich (Rechnung/Prognose).
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) IT-Expert-Query/1.0"
    }
    try:
        response = requests.get(
            CH_EFV_MAIN_CSV, headers=headers, timeout=25
        )
        response.raise_for_status()
        reader = csv.DictReader(io.StringIO(response.text))
        rows = list(reader)
        if not rows:
            print("Schweiz: Keine Daten in der CSV.")
            return None
        # hh="staat" = Konföderation/Bund, variable "saldo" = Finanzierungssaldo, "bruttoschuld_fs" = Bruttoschuld
        staat = [r for r in rows if (r.get("hh") or "").strip() == "staat"]
        saldo_rows = [r for r in staat if (r.get("variable") or "").strip() == "saldo"]
        schuld_rows = [r for r in staat if (r.get("variable") or "").strip() == "bruttoschuld_fs"]
        if not saldo_rows:
            print("Schweiz: Kein Finanzierungssaldo (variable=saldo) gefunden.")
            return None
        latest_saldo = max(saldo_rows, key=lambda r: int((r.get("jahr") or "0").strip() or "0"))
        jahr = latest_saldo.get("jahr", "").strip()
        try:
            saldo_mio = float((latest_saldo.get("value") or "0").replace(" ", ""))
        except ValueError:
            print("Schweiz: Saldo-Wert konnte nicht gelesen werden.")
            return None
        source = (latest_saldo.get("source") or "").strip()
        # Optional: Bruttoschuld für dasselbe Jahr
        schuld_mio = None
        for r in schuld_rows:
            if (r.get("jahr") or "").strip() == jahr:
                try:
                    schuld_mio = float((r.get("value") or "0").replace(" ", ""))
                except ValueError:
                    pass
                break
        print("--- Offizieller Finanzstatus Bund Schweiz (EFV) ---")
        print(f"Jahr:  {jahr}" + (f"  ({source})" if source else ""))
        print(f"Finanzierungssaldo (Bund):  {saldo_mio:,.2f} Mio. CHF")
        if saldo_mio >= 0:
            print(f"  (= Überschuss {saldo_mio / 1_000:,.2f} Mrd. CHF)")
        else:
            print(f"  (= Defizit {abs(saldo_mio) / 1_000:,.2f} Mrd. CHF)")
        if schuld_mio is not None:
            print(f"Bruttoschuld (Bund):  {schuld_mio:,.2f} Mio. CHF  (= {schuld_mio / 1_000:,.2f} Mrd. CHF)")
        return {
            "country": "ch",
            "date": jahr,
            "value": saldo_mio / 1_000,
            "value_mio": saldo_mio,
            "unit": "Mrd. CHF",
            "label": "Finanzierungssaldo (Bund)",
            "bruttoschuld_mio": schuld_mio,
        }
    except requests.exceptions.RequestException as e:
        print(f"Netzwerk- oder API-Fehler: {e}")
        return None
    except Exception as e:
        print(f"Fehler beim Verarbeiten der Daten: {e}")
        return None


def get_li_account_balance():
    """
    Liechtenstein: Derzeit keine öffentliche API oder stabile CSV/JSON-URL
    für den Staatshaushalt oder Kassenstand. Das Open-Data-Portal opendata.li
    (seit Dez. 2024) enthält noch keine maschinenlesbaren Haushaltsdaten;
    Angaben erscheinen in PDFs (z. B. Wirtschafts- und Finanzdaten, Rechenschaftsbericht).
    """
    print("--- Liechtenstein: Keine programmatische Abfrage möglich ---")
    print()
    print("Für den Staatshaushalt bzw. Kassenstand stellt Liechtenstein derzeit")
    print("keine oeffentliche API und keine stabile CSV/JSON-URL bereit.")
    print()
    print("Weitere Informationen:")
    print("  • Open Data Portal:  https://www.opendata.li/")
    print("  • Regierung / Finanzen:  https://www.regierung.li/")
    print("  • Landesverwaltung (Rechenschaftsbericht, Voranschlag):  https://www.llv.li/")
    print()
    print("Haushalts- und Finanzzahlen werden in PDF-Publikationen veroeffentlicht")
    print("(z. B. Wirtschafts- und Finanzdaten, Rechenschaftsbericht, Landesvoranschlag).")


if __name__ == "__main__":
    import sys
    argv = [a for a in sys.argv[1:] if a != "--save"]
    do_save = "--save" in sys.argv
    arg = argv[0].lower() if argv else ""
    result = None
    if arg in ("de", "deutschland", "germany"):
        result = get_de_account_balance()
    elif arg in ("at", "österreich", "austria", "oesterreich"):
        result = get_at_account_balance()
    elif arg in ("ca", "kanada", "canada"):
        result = get_ca_account_balance()
    elif arg in ("mx", "mexiko", "mexico"):
        result = get_mx_account_balance()
    elif arg in ("ch", "schweiz", "switzerland"):
        result = get_ch_account_balance()
    elif arg in ("li", "liechtenstein"):
        get_li_account_balance()
    else:
        result = get_us_account_balance_pro()
    if do_save and result and isinstance(result, dict):
        import persist
        extra = {k: v for k, v in result.items() if k not in ("country", "date", "value", "unit", "label")}
        persist.save_record(
            country=result.get("country", ""),
            date=result.get("date", ""),
            value=result.get("value", 0),
            unit=result.get("unit", ""),
            label=result.get("label", ""),
            **extra,
        )
        print(f"\n[Daten gespeichert: {persist.DB_PATH}]")